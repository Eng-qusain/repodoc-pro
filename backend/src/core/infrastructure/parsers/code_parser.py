"""
File Parsers — CSV, Excel, Image, and Code file processors.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


# ─── CSV Parser ───────────────────────────────────────────────────────────────

class CSVParser:
    """Parse CSV files and extract preview data and statistics."""

    def parse(self, file_path: str, max_rows: int = 100) -> dict:
        try:
            import pandas as pd

            df = pd.read_csv(file_path, nrows=max_rows + 1, on_bad_lines="skip")
            truncated = len(df) > max_rows
            preview_df = df.head(max_rows)

            # Column stats
            col_types = {col: str(dtype) for col, dtype in df.dtypes.items()}

            # Full row count (without loading full file)
            full_rows = self._count_rows(file_path)

            # Descriptive stats for numeric columns
            desc = {}
            for col in df.select_dtypes(include=["number"]).columns[:10]:
                desc[col] = {
                    "min": float(df[col].min()),
                    "max": float(df[col].max()),
                    "mean": float(df[col].mean()),
                    "null_count": int(df[col].isnull().sum()),
                }

            # Convert to safe JSON rows (handle NaN, NaT, etc.)
            rows = []
            for _, row in preview_df.iterrows():
                rows.append([str(v) if v is not None else "" for v in row])

            return {
                "headers": list(df.columns),
                "rows": rows,
                "stats": {
                    "row_count": full_rows,
                    "column_count": len(df.columns),
                    "column_types": col_types,
                    "numeric_stats": desc,
                    "null_counts": {col: int(df[col].isnull().sum()) for col in df.columns},
                    "truncated": truncated,
                },
            }
        except ImportError:
            return self._parse_stdlib(file_path, max_rows)
        except Exception as e:
            logger.warning(f"CSV parse error {file_path}: {e}")
            return {"headers": [], "rows": [], "stats": {"error": str(e)}}

    @staticmethod
    def _count_rows(file_path: str) -> int:
        try:
            with open(file_path, "rb") as f:
                return sum(1 for _ in f) - 1  # subtract header
        except Exception:
            return 0

    def _parse_stdlib(self, file_path: str, max_rows: int) -> dict:
        """Fallback CSV parser using stdlib csv module."""
        import csv
        headers: list[str] = []
        rows: list[list[str]] = []
        try:
            with open(file_path, newline="", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                headers = next(reader, [])
                for i, row in enumerate(reader):
                    if i >= max_rows:
                        break
                    rows.append(row)
        except Exception as e:
            return {"headers": [], "rows": [], "stats": {"error": str(e)}}

        return {
            "headers": headers,
            "rows": rows,
            "stats": {
                "row_count": len(rows),
                "column_count": len(headers),
                "column_types": {h: "string" for h in headers},
            },
        }


# ─── Excel Parser ─────────────────────────────────────────────────────────────

class ExcelParser:
    """Parse Excel (.xlsx / .xls) workbooks."""

    def parse(self, file_path: str, max_rows: int = 50) -> dict:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                headers = []

                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        headers = [str(c) if c is not None else "" for c in row]
                    elif i <= max_rows:
                        rows.append([str(c) if c is not None else "" for c in row])
                    else:
                        break

                sheets.append({
                    "name": sheet_name,
                    "headers": headers,
                    "rows": rows,
                    "row_count": ws.max_row,
                    "col_count": ws.max_column,
                })

            wb.close()
            return {
                "workbook_name": Path(file_path).name,
                "sheet_count": len(sheets),
                "sheets": sheets,
            }

        except ImportError:
            return {"error": "openpyxl not installed", "sheets": []}
        except Exception as e:
            logger.warning(f"Excel parse error {file_path}: {e}")
            return {"error": str(e), "sheets": []}


# ─── Image Parser ─────────────────────────────────────────────────────────────

class ImageParser:
    """Extract metadata from image files."""

    def parse(self, file_path: str) -> dict:
        info: dict[str, Any] = {"path": file_path, "name": Path(file_path).name}

        try:
            from PIL import Image as PILImage
            with PILImage.open(file_path) as img:
                info["width"] = img.width
                info["height"] = img.height
                info["mode"] = img.mode
                info["format"] = img.format
                if hasattr(img, "_getexif") and img._getexif():
                    info["has_exif"] = True
        except ImportError:
            pass
        except Exception as e:
            info["error"] = str(e)

        return info


# ─── Code Parser ─────────────────────────────────────────────────────────────

class CodeParser:
    """
    Parse and analyze source code files.
    Extracts functions, classes, imports using AST (Python)
    or regex for other languages.
    """

    def parse(self, file_path: str, content: str, language: str) -> dict:
        result: dict = {
            "language": language,
            "line_count": content.count("\n") + 1,
            "functions": [],
            "classes": [],
            "imports": [],
            "complexity_score": 0,
        }

        if language.lower() == "python":
            result.update(self._parse_python(content))
        elif language.lower() in ("javascript", "typescript", "react (tsx)"):
            result.update(self._parse_js_ts(content))

        return result

    @staticmethod
    def _parse_python(content: str) -> dict:
        """Use Python AST to extract structure."""
        import ast
        try:
            tree = ast.parse(content)
        except SyntaxError:
            return {}

        functions, classes, imports = [], [], []

        for node in ast.walk(tree):
            if isinstance(node, ast.FunctionDef | ast.AsyncFunctionDef):
                args = [a.arg for a in node.args.args]
                functions.append({
                    "name": node.name,
                    "line": node.lineno,
                    "args": args,
                    "is_async": isinstance(node, ast.AsyncFunctionDef),
                    "docstring": ast.get_docstring(node) or "",
                })
            elif isinstance(node, ast.ClassDef):
                methods = [
                    n.name for n in ast.walk(node)
                    if isinstance(n, ast.FunctionDef | ast.AsyncFunctionDef)
                ]
                classes.append({
                    "name": node.name,
                    "line": node.lineno,
                    "methods": methods,
                    "docstring": ast.get_docstring(node) or "",
                })
            elif isinstance(node, ast.Import):
                for alias in node.names:
                    imports.append(alias.name)
            elif isinstance(node, ast.ImportFrom):
                imports.append(node.module or "")

        # Cyclomatic complexity approximation
        complexity = sum(
            1 for node in ast.walk(tree)
            if isinstance(node, (
                ast.If, ast.For, ast.While, ast.Try,
                ast.ExceptHandler, ast.With, ast.Assert,
            ))
        )

        return {
            "functions": functions,
            "classes": classes,
            "imports": list(set(filter(None, imports))),
            "complexity_score": complexity,
        }

    @staticmethod
    def _parse_js_ts(content: str) -> dict:
        """Regex-based JS/TS structure extraction."""
        import re

        functions = re.findall(
            r"(?:async\s+)?function\s+(\w+)\s*\(|const\s+(\w+)\s*=\s*(?:async\s+)?\(",
            content,
        )
        classes = re.findall(r"class\s+(\w+)", content)

        # Handles: import X from '...', import { X } from '...', import type X from '...',
        # import '...', require('...')
        import_pattern = re.compile(
            r"""(?x)
                (?:
                    import\s+(?:type\s+)?
                    (?:[^'";\n]*?\s+from\s+)?
                    ['"]([^'"]+)['"]
                )
                |
                (?:require\s*\(\s*['"]([^'"]+)['"]\s*\))
            """,
            re.MULTILINE,
        )
        raw = import_pattern.findall(content)
        imports = list(set(m[0] or m[1] for m in raw if m[0] or m[1]))

        func_names = [f[0] or f[1] for f in functions if f[0] or f[1]]
        return {
            "functions": [{"name": n, "line": 0} for n in func_names],
            "classes": [{"name": c, "line": 0} for c in classes],
            "imports": list(set(imports)),
        }
    